#!/usr/bin/env python3
"""JSON スペックから Excel (.xlsx) 帳票を生成するビルダー。

使い方:
    uv run python scripts/xlsx_builder.py build --spec spec.json
    cat spec.json | uv run python scripts/xlsx_builder.py build
    uv run python scripts/xlsx_builder.py example   # サンプル spec を標準出力

スペックの詳細は references/spec.md を参照。
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_HEADER_STYLE = {"bold": True, "bg": "4472C4", "font_color": "FFFFFF"}
CHART_TYPES = {"bar": BarChart, "line": LineChart, "pie": PieChart}


def _maybe_date(value: Any, number_format: str | None):
    """number_format が日付系で値が ISO 文字列ならば date/datetime に変換する。"""
    if not isinstance(value, str) or not number_format:
        return value
    fmt = number_format.lower()
    if "y" not in fmt and "d" not in fmt:
        return value
    for parser in (dt.date.fromisoformat, dt.datetime.fromisoformat):
        try:
            return parser(value)
        except ValueError:
            continue
    return value


def _apply_header_style(cell, style: dict) -> None:
    cell.font = Font(bold=style.get("bold", True), color=style.get("font_color", "FFFFFF"))
    if style.get("bg"):
        cell.fill = PatternFill("solid", fgColor=style["bg"])
    cell.alignment = Alignment(horizontal=style.get("align", "center"), vertical="center")


def _build_sheet(ws, sheet: dict) -> None:
    columns = sheet["columns"]
    keys = [c["key"] for c in columns]
    n_cols = len(columns)

    # --- ヘッダー行 ---
    header_style = {**DEFAULT_HEADER_STYLE, **sheet.get("header_style", {})}
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col.get("header", col["key"]))
        _apply_header_style(cell, header_style)
        if col.get("width"):
            ws.column_dimensions[get_column_letter(col_idx)].width = col["width"]

    # --- データ行 ---
    rows = sheet.get("rows", [])
    for r_off, row in enumerate(rows):
        excel_row = 2 + r_off
        for col_idx, col in enumerate(columns, start=1):
            raw = row.get(col["key"])
            value = _maybe_date(raw, col.get("number_format"))
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            if col.get("number_format"):
                cell.number_format = col["number_format"]
            if col.get("align"):
                cell.alignment = Alignment(horizontal=col["align"])

    last_data_row = 1 + len(rows)

    # --- 合計行 ---
    total = sheet.get("total_row")
    if total and rows:
        total_row = last_data_row + 1
        label_idx = keys.index(total["label_col"]) + 1 if total.get("label_col") in keys else 1
        lc = ws.cell(row=total_row, column=label_idx, value=total.get("label", "合計"))
        lc.font = Font(bold=True)
        for sum_key in total.get("sums", []):
            if sum_key not in keys:
                continue
            cidx = keys.index(sum_key) + 1
            letter = get_column_letter(cidx)
            sc = ws.cell(row=total_row, column=cidx,
                         value=f"=SUM({letter}2:{letter}{last_data_row})")
            sc.font = Font(bold=True)
            col = columns[cidx - 1]
            if col.get("number_format"):
                sc.number_format = col["number_format"]

    # --- フリーズペイン / オートフィルタ ---
    if sheet.get("freeze"):
        ws.freeze_panes = sheet["freeze"]
    if sheet.get("auto_filter") and rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{last_data_row}"

    # --- 条件付き書式 ---
    for cond in sheet.get("conditional", []):
        if cond["range_col"] not in keys or not rows:
            continue
        letter = get_column_letter(keys.index(cond["range_col"]) + 1)
        rng = f"{letter}2:{letter}{last_data_row}"
        ctype = cond.get("type", "color_scale")
        if ctype == "color_scale":
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="min", start_color=cond.get("min_color", "F8696B"),
                mid_type="percentile", mid_value=50, mid_color=cond.get("mid_color", "FFEB84"),
                end_type="max", end_color=cond.get("max_color", "63BE7B")))
        elif ctype == "data_bar":
            ws.conditional_formatting.add(rng, DataBarRule(
                start_type="min", end_type="max", color=cond.get("color", "638EC6")))
        elif ctype == "greater_than":
            ws.conditional_formatting.add(rng, CellIsRule(
                operator="greaterThan", formula=[str(cond["value"])],
                fill=PatternFill("solid", fgColor=cond.get("fill", "FFC7CE"))))

    # --- グラフ ---
    for chart_spec in sheet.get("charts", []):
        if not rows:
            continue
        cls = CHART_TYPES.get(chart_spec.get("type", "bar"))
        if cls is None:
            continue
        chart = cls()
        chart.title = chart_spec.get("title")
        cat_idx = keys.index(chart_spec["categories_col"]) + 1
        cats = Reference(ws, min_col=cat_idx, min_row=2, max_row=last_data_row)
        value_cols = chart_spec.get("values_cols") or [chart_spec["values_col"]]
        for vkey in value_cols:
            vidx = keys.index(vkey) + 1
            data = Reference(ws, min_col=vidx, min_row=1, max_row=last_data_row)
            chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, chart_spec.get("anchor", "H2"))


def build(spec: dict) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    props = spec.get("properties", {})
    if props.get("title"):
        wb.properties.title = props["title"]
    if props.get("creator"):
        wb.properties.creator = props["creator"]

    sheets = spec.get("sheets", [])
    if not sheets:
        raise ValueError("spec.sheets が空です。少なくとも 1 シート必要です。")
    for sheet in sheets:
        ws = wb.create_sheet(title=sheet["name"][:31])
        _build_sheet(ws, sheet)

    filename = spec.get("filename", "report.xlsx")
    wb.save(filename)
    return filename


EXAMPLE_SPEC = {
    "filename": "sales-report.xlsx",
    "properties": {"title": "月次売上レポート", "creator": "xlsx-report-builder"},
    "sheets": [
        {
            "name": "売上明細",
            "freeze": "A2",
            "auto_filter": True,
            "header_style": {"bg": "305496", "font_color": "FFFFFF"},
            "columns": [
                {"header": "日付", "key": "date", "width": 14, "number_format": "yyyy-mm-dd"},
                {"header": "商品", "key": "product", "width": 18},
                {"header": "数量", "key": "qty", "width": 10, "number_format": "#,##0", "align": "right"},
                {"header": "金額", "key": "amount", "width": 14, "number_format": "¥#,##0", "align": "right"},
            ],
            "rows": [
                {"date": "2026-05-01", "product": "Aプラン", "qty": 3, "amount": 30000},
                {"date": "2026-05-03", "product": "Bプラン", "qty": 1, "amount": 8000},
                {"date": "2026-05-07", "product": "Aプラン", "qty": 5, "amount": 50000},
                {"date": "2026-05-12", "product": "Cプラン", "qty": 2, "amount": 24000},
            ],
            "total_row": {"label_col": "product", "label": "合計", "sums": ["qty", "amount"]},
            "conditional": [{"range_col": "amount", "type": "data_bar", "color": "638EC6"}],
            "charts": [{
                "type": "bar", "title": "商品別 金額",
                "categories_col": "product", "values_col": "amount", "anchor": "G2",
            }],
        }
    ],
}


def main() -> int:
    parser = argparse.ArgumentParser(description="JSON スペックから .xlsx 帳票を生成する")
    sub = parser.add_subparsers(dest="command", required=True)
    b = sub.add_parser("build", help="spec から xlsx を生成")
    b.add_argument("--spec", help="スペック JSON ファイル（省略時は stdin）")
    sub.add_parser("example", help="サンプル spec を標準出力に表示")

    args = parser.parse_args()

    if args.command == "example":
        json.dump(EXAMPLE_SPEC, sys.stdout, ensure_ascii=False, indent=2)
        print()
        return 0

    if args.spec:
        with open(args.spec, encoding="utf-8") as f:
            spec = json.load(f)
    else:
        spec = json.load(sys.stdin)

    filename = build(spec)
    print(f"生成しました: {filename}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
