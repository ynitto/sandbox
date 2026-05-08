"""Raw content ingestion for PDF and Excel files."""
from __future__ import annotations
from pathlib import Path
from typing import Iterator
from dataclasses import dataclass

import openpyxl
from PIL import Image


@dataclass
class PageImage:
    page: int
    image: Image.Image


@dataclass
class ExcelSheet:
    name: str
    rows: list[list[str]]  # raw cell values as strings


def iter_pdf_images(path: Path, dpi: int = 150) -> Iterator[PageImage]:
    """Render each PDF page to a PIL Image using pypdfium2 (no system deps)."""
    import pypdfium2 as pdfium

    doc = pdfium.PdfDocument(str(path))
    scale = dpi / 72.0
    for page_num in range(len(doc)):
        page = doc[page_num]
        bitmap = page.render(scale=scale, rotation=0)
        pil_img = bitmap.to_pil()
        yield PageImage(page=page_num, image=pil_img)
    doc.close()


def iter_excel_sheets(path: Path) -> Iterator[ExcelSheet]:
    """Yield each sheet as a list of rows (cell values as strings)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            str_row = [str(v) if v is not None else "" for v in row]
            if any(str_row):  # skip fully empty rows
                rows.append(str_row)
        if rows:
            yield ExcelSheet(name=name, rows=rows)
    wb.close()


def pdf_text_by_page(path: Path) -> dict[int, str]:
    """Extract plain text per page using pdfplumber (fallback / metadata)."""
    try:
        import pdfplumber
        result: dict[int, str] = {}
        with pdfplumber.open(path) as pdf:
            for i, page in enumerate(pdf.pages):
                result[i] = page.extract_text() or ""
        return result
    except Exception:
        return {}
